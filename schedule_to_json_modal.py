import argparse
import json
import re
from datetime import date, datetime, timedelta
from io import BytesIO
from pathlib import Path
from typing import List, Optional, Sequence, Tuple

from openpyxl import load_workbook

try:
    import modal
except ImportError:  # Allows plain-python local use even when Modal is not installed.
    modal = None


MONTH_MAP = {
    "JAN": 1,
    "JANUARY": 1,
    "FEB": 2,
    "FEBRUARY": 2,
    "MAR": 3,
    "MARCH": 3,
    "APR": 4,
    "APRIL": 4,
    "MAY": 5,
    "JUN": 6,
    "JUNE": 6,
    "JUL": 7,
    "JULY": 7,
    "AUG": 8,
    "AUGUST": 8,
    "SEP": 9,
    "SEPT": 9,
    "SEPTEMBER": 9,
    "OCT": 10,
    "OCTOBER": 10,
    "NOV": 11,
    "NOVEMBER": 11,
    "DEC": 12,
    "DECEMBER": 12,
}

SHIFT_MAP = {
    "D": ("Day Shift", "on_unit"),
    "N": ("Night Shift", "on_unit"),
    "OT": ("Overtime", "on_unit"),
    "ED": ("Education Day", "off_unit"),
    "UC": ("Unit Council", "off_unit"),
    "ILL": ("Sick Day", "off"),
    "VAC": ("Vacation", "off"),
    "SH": ("Stat Holiday", "off"),
    "FAMILY": ("Family Emergency", "off"),
}


def clean_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_name(raw_name: str) -> str:
    text = clean_text(raw_name).upper()
    text = re.sub(r"\s*-\s*.*$", "", text)  # remove trailing annotations like "- MAT LEAVE"
    text = re.sub(r"\s*\([^)]*\)", "", text)  # remove parenthetical notes like "(1)"
    text = re.sub(r"\s+", " ", text).strip()
    if "," in text:
        last, first = [part.strip() for part in text.split(",", 1)]
        return re.sub(r"\s+", " ", f"{first} {last}").strip()
    return text


def find_header_rows(ws) -> Tuple[int, int]:
    """Find the month-label row and the day-number row immediately below it."""
    max_scan_rows = min(ws.max_row, 30)
    for row in range(1, max_scan_rows):
        month_hits = 0
        for col in range(1, ws.max_column + 1):
            value = clean_text(ws.cell(row, col).value).upper()
            if value in MONTH_MAP:
                month_hits += 1
        if month_hits >= 2:
            next_row = row + 1
            day_hits = 0
            for col in range(1, ws.max_column + 1):
                value = ws.cell(next_row, col).value
                if isinstance(value, (int, float)) and 1 <= int(value) <= 31:
                    day_hits += 1
            if day_hits >= 7:
                return row, next_row
    raise ValueError("Could not detect the month/day header rows.")


def find_name_column(ws, data_start_row: int) -> int:
    """Find the column that most likely contains staff names (e.g. LAST, FIRST)."""
    best_col = None
    best_score = -1
    for col in range(1, min(ws.max_column, 10) + 1):
        score = 0
        for row in range(data_start_row, ws.max_row + 1):
            value = clean_text(ws.cell(row, col).value)
            if "," in value:
                score += 1
        if score > best_score:
            best_col = col
            best_score = score
    if not best_col or best_score <= 0:
        raise ValueError("Could not detect the staff-name column.")
    return best_col


def detect_date_columns(ws, day_row: int, name_col: int) -> List[int]:
    """Find schedule date columns by scanning right of the name column for day numbers."""
    date_cols: List[int] = []
    for col in range(name_col + 1, ws.max_column + 1):
        value = ws.cell(day_row, col).value
        if isinstance(value, (int, float)) and 1 <= int(value) <= 31:
            date_cols.append(col)
    if not date_cols:
        raise ValueError("Could not detect the schedule date columns.")
    return date_cols


def first_visible_month(month_row_values: Sequence[str]) -> int:
    for raw in month_row_values:
        label = clean_text(raw).upper()
        if label in MONTH_MAP:
            return MONTH_MAP[label]
    raise ValueError("Could not detect the first visible month label.")


def infer_start_date(wb, ws, month_row: int, day_row: int, date_cols: List[int]) -> date:
    """
    Infer the schedule start date.

    This schedule template visually groups dates by week, so month labels may only appear
    once per 7-day block. The most reliable strategy is:
    1) take the first visible month label,
    2) take the first day number in the schedule,
    3) choose the year closest to the workbook modified/created date.
    """
    start_month = first_visible_month(ws.cell(month_row, col).value for col in date_cols)
    start_day = int(ws.cell(day_row, date_cols[0]).value)

    modified_dt = getattr(wb.properties, "modified", None) or getattr(wb.properties, "created", None)
    if modified_dt is None:
        anchor_date = datetime.today().date()
    else:
        anchor_date = modified_dt.date() if hasattr(modified_dt, "date") else modified_dt

    candidates = []
    for year in {anchor_date.year - 1, anchor_date.year, anchor_date.year + 1}:
        start = date(year, start_month, start_day)
        midpoint = start + timedelta(days=max(len(date_cols) - 1, 0) // 2)
        score = abs((anchor_date - midpoint).days)
        candidates.append((score, start))

    return min(candidates, key=lambda item: item[0])[1]


def parse_schedule_records_from_workbook(wb, sheet_name: Optional[str] = None) -> List[dict]:
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]

    month_row, day_row = find_header_rows(ws)
    data_start_row = day_row + 1
    name_col = find_name_column(ws, data_start_row)
    date_cols = detect_date_columns(ws, day_row, name_col)
    start_date = infer_start_date(wb, ws, month_row, day_row, date_cols)
    dates_by_col = {col: (start_date + timedelta(days=idx)).isoformat() for idx, col in enumerate(date_cols)}

    records: List[dict] = []
    for row in range(data_start_row, ws.max_row + 1):
        raw_name = clean_text(ws.cell(row, name_col).value)
        if not raw_name or "," not in raw_name:
            continue

        full_name = normalize_name(raw_name)
        for col in date_cols:
            raw_shift = clean_text(ws.cell(row, col).value)
            if not raw_shift:
                continue
            shift_key = raw_shift.upper()
            if shift_key not in SHIFT_MAP:
                continue

            description, status = SHIFT_MAP[shift_key]
            records.append(
                {
                    "date": dates_by_col[col],
                    "name": full_name,
                    "shift_type": raw_shift,
                    "description": description,
                    "status": status,
                }
            )

    return records


def parse_schedule_bytes(file_bytes: bytes, sheet_name: Optional[str] = None) -> List[dict]:
    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    return parse_schedule_records_from_workbook(wb, sheet_name)


def convert_file_to_json(input_xlsx: str, output_json: str = "output.json", sheet_name: Optional[str] = None) -> List[dict]:
    wb = load_workbook(input_xlsx, data_only=True)
    records = parse_schedule_records_from_workbook(wb, sheet_name)
    with open(output_json, "w", encoding="utf-8") as f:
        json.dump(records, f, ensure_ascii=False, indent=2)
    return records


if modal is not None:
    app = modal.App("schedule-to-json-converter")
    image = modal.Image.debian_slim(python_version="3.11").pip_install("openpyxl")

    @app.function(image=image, timeout=300)
    def convert_schedule_remote(file_bytes: bytes, sheet_name: Optional[str] = None) -> str:
        records = parse_schedule_bytes(file_bytes, sheet_name)
        return json.dumps(records, ensure_ascii=False, indent=2)


    @app.local_entrypoint()
    def main(input_xlsx: str, output_json: str = "output.json", sheet_name: str = ""):
        with open(input_xlsx, "rb") as f:
            file_bytes = f.read()

        json_text = convert_schedule_remote.remote(file_bytes, sheet_name or None)
        Path(output_json).write_text(json_text, encoding="utf-8")
        print(f"Wrote JSON to {Path(output_json).resolve()}")


def _plain_python_cli():
    parser = argparse.ArgumentParser(
        description="Convert a nursing schedule Excel workbook into normalized JSON shift records."
    )
    parser.add_argument("input_xlsx", help="Path to the input Excel workbook (.xlsx)")
    parser.add_argument(
        "output_json",
        nargs="?",
        default="output.json",
        help="Path to the output JSON file (default: output.json)",
    )
    parser.add_argument(
        "--sheet",
        dest="sheet_name",
        default=None,
        help="Optional worksheet name. Defaults to the first worksheet.",
    )
    args = parser.parse_args()
    records = convert_file_to_json(args.input_xlsx, args.output_json, args.sheet_name)
    print(f"Wrote {len(records)} records to {Path(args.output_json).resolve()}")


if __name__ == "__main__":
    _plain_python_cli()
