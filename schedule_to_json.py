import argparse
import json
import re
from datetime import date, datetime, timedelta
from io import BytesIO
from pathlib import Path
from typing import List, Optional, Sequence, Tuple

from openpyxl import load_workbook


MONTH_MAP = {
    "JAN": 1, "JANUARY": 1,
    "FEB": 2, "FEBRUARY": 2,
    "MAR": 3, "MARCH": 3,
    "APR": 4, "APRIL": 4,
    "MAY": 5,
    "JUN": 6, "JUNE": 6,
    "JUL": 7, "JULY": 7,
    "AUG": 8, "AUGUST": 8,
    "SEP": 9, "SEPT": 9, "SEPTEMBER": 9,
    "OCT": 10, "OCTOBER": 10,
    "NOV": 11, "NOVEMBER": 11,
    "DEC": 12, "DECEMBER": 12,
}

SHIFT_MAP = {
    "D": ("Day Shift", "on_unit"),
    "N": ("Night Shift", "on_unit"),
    "OT": ("Overtime", "on_unit"),
    "ED": ("Education Day", "off_unit"),
    "UC": ("Unit Council", "off_unit"),
    "ILL": ("Sick Day", "off"),
    "VAC": ("Vacation", "off"),
    "SH": ("Stat Holiday", "off"),
    "FAMILY": ("Family Emergency", "off"),
}


def clean_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_name(raw_name: str) -> str:
    text = clean_text(raw_name).upper()
    text = re.sub(r"\s*-\s*.*$", "", text)
    text = re.sub(r"\s*\([^)]*\)", "", text)
    text = re.sub(r"\s+", " ", text).strip()
    if "," in text:
        last, first = [part.strip() for part in text.split(",", 1)]
        return f"{first} {last}".strip()
    return text


def find_header_rows(ws) -> Tuple[int, int]:
    max_scan_rows = min(ws.max_row, 30)
    for row in range(1, max_scan_rows):
        month_hits = 0
        for col in range(1, ws.max_column + 1):
            value = clean_text(ws.cell(row, col).value).upper()
            if value in MONTH_MAP:
                month_hits += 1
        if month_hits >= 2:
            next_row = row + 1
            day_hits = 0
            for col in range(1, ws.max_column + 1):
                value = ws.cell(next_row, col).value
                if isinstance(value, (int, float)) and 1 <= int(value) <= 31:
                    day_hits += 1
            if day_hits >= 7:
                return row, next_row
    raise ValueError("Could not detect header rows")


def find_name_column(ws, start_row: int) -> int:
    best_col = None
    best_score = -1
    for col in range(1, 11):
        score = 0
        for row in range(start_row, ws.max_row + 1):
            value = clean_text(ws.cell(row, col).value)
            if "," in value:
                score += 1
        if score > best_score:
            best_col = col
            best_score = score
    if not best_col:
        raise ValueError("Name column not found")
    return best_col


def detect_date_columns(ws, day_row: int, name_col: int) -> List[int]:
    cols = []
    for col in range(name_col + 1, ws.max_column + 1):
        v = ws.cell(day_row, col).value
        if isinstance(v, (int, float)) and 1 <= int(v) <= 31:
            cols.append(col)
    return cols


def infer_start_date(wb, ws, month_row, day_row, date_cols):
    modified = getattr(wb.properties, "modified", None) or datetime.today()
    base_year = modified.year
    month = MONTH_MAP[clean_text(ws.cell(month_row, date_cols[0]).value).upper()]
    day = int(ws.cell(day_row, date_cols[0]).value)
    return date(base_year, month, day)


def parse_schedule_records_from_workbook(wb):
    ws = wb[wb.sheetnames[0]]

    month_row, day_row = find_header_rows(ws)
    name_col = find_name_column(ws, day_row + 1)
    date_cols = detect_date_columns(ws, day_row, name_col)
    start_date = infer_start_date(wb, ws, month_row, day_row, date_cols)

    records = []

    for row in range(day_row + 1, ws.max_row + 1):
        raw_name = clean_text(ws.cell(row, name_col).value)
        if not raw_name or "," not in raw_name:
            continue

        name = normalize_name(raw_name)

        for i, col in enumerate(date_cols):
            shift = clean_text(ws.cell(row, col).value)
            if not shift:
                continue

            key = shift.upper()
            if key not in SHIFT_MAP:
                continue

            desc, status = SHIFT_MAP[key]

            records.append({
                "date": (start_date + timedelta(days=i)).isoformat(),
                "name": name,
                "shift_type": shift,
                "description": desc,
                "status": status,
            })

    return records


def parse_schedule(input_path, output_path="output.json"):
    wb = load_workbook(input_path, data_only=True)
    records = parse_schedule_records_from_workbook(wb)

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(records, f, indent=2)

    return records
